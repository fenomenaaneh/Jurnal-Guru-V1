from http.server import BaseHTTPRequestHandler
import json, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    "indigo":  "3730A3", "indigo2": "4F46E5", "indigo3": "6366F1",
    "indigoL": "EEF2FF", "white":   "FFFFFF", "slate50": "F8FAFC",
    "slate100":"F1F5F9", "slate200":"E2E8F0",  "dark":   "1E293B",
    "emerald": "059669", "emeraldL":"D1FAE5",  "amber":  "D97706",
    "amberL":  "FEF3C7", "blue":    "2563EB",  "blueL":  "DBEAFE",
    "rose":    "E11D48", "roseL":   "FFE4E6",  "gray":   "94A3B8",
}
STATUS_FG = {"H":C["emerald"],"S":C["amber"],"I":C["blue"],"A":C["rose"],"-":C["gray"]}
STATUS_BG = {"H":C["emeraldL"],"S":C["amberL"],"I":C["blueL"],"A":C["roseL"],"-":C["slate100"]}

def fill(color): return PatternFill('solid', start_color=color, end_color=color)
def thin(color=None):
    s = Side(style='thin', color=color or C["slate200"])
    return Border(top=s, bottom=s, left=s, right=s)
def fnt(bold=False, size=9, color=None):
    return Font(name='Arial', bold=bold, size=size, color=color or C["dark"])
def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def merge(ws, r1, c1, r2, c2):
    # WAJIB pakai keyword args — positional args deprecated di openpyxl terbaru
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def sc(ws, r, c, val='', bold=False, size=9, fg=None, bg=None, h='left', bdr=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = fnt(bold=bold, size=size, color=fg)
    cell.alignment = aln(h=h)
    if bg:  cell.fill   = fill(bg)
    if bdr: cell.border = thin()
    return cell

def build_excel(data: dict) -> bytes:
    school    = data.get("school", "SMPN 21 Jambi")
    mapel     = data.get("mapel", "")
    kelas     = data.get("kelas", "")
    guru      = data.get("guru", "")
    dicetak   = data.get("dicetak", "")
    pertemuan = data.get("pertemuan", [])
    siswa     = data.get("siswa", [])

    N  = len(pertemuan)
    NS = len(siswa)

    COL_NO=1; COL_NIS=2; COL_NAMA=3
    COL_P1=4; COL_PE=COL_P1+N-1
    COL_H=COL_PE+1; COL_S=COL_PE+2; COL_I=COL_PE+3
    COL_A=COL_PE+4; COL_TOT=COL_PE+5; COL_PCT=COL_PE+6
    LAST=COL_PCT

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kehadiran"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions[get_column_letter(COL_NO)].width   = 5
    ws.column_dimensions[get_column_letter(COL_NIS)].width  = 13
    ws.column_dimensions[get_column_letter(COL_NAMA)].width = 30
    for i in range(N):
        ws.column_dimensions[get_column_letter(COL_P1+i)].width = 8
    for ci in [COL_H, COL_S, COL_I, COL_A]:
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions[get_column_letter(COL_TOT)].width = 8
    ws.column_dimensions[get_column_letter(COL_PCT)].width = 10

    # Baris 1: Banner sekolah
    ws.row_dimensions[1].height = 32
    merge(ws, 1,1,1,LAST)
    c=ws.cell(row=1, column=1, value=school)
    c.font=fnt(True,15,C["white"]); c.fill=fill(C["indigo"]); c.alignment=aln('center')

    # Baris 2: Judul
    ws.row_dimensions[2].height = 24
    merge(ws, 2,1,2,LAST)
    c=ws.cell(row=2, column=1, value="REKAP KEHADIRAN SISWA")
    c.font=fnt(True,12,C["indigo"]); c.fill=fill(C["indigoL"]); c.alignment=aln('center')

    # Baris 3: spasi
    ws.row_dimensions[3].height = 5

    # Baris 4-6: Info
    half = LAST // 2
    info = [
        ("Mata Pelajaran", mapel,   "Kelas",            kelas),
        ("Guru",           guru,    "Jumlah Pertemuan", str(N)),
        ("Dicetak",        dicetak, "",                 ""),
    ]
    for ri, (k1, v1, k2, v2) in enumerate(info):
        r = 4 + ri
        ws.row_dimensions[r].height = 17
        merge(ws, r,1,r,half)
        cl = ws.cell(row=r, column=1, value=f"  {k1}  :  {v1}")
        cl.font=fnt(ri==0, 9); cl.fill=fill(C["slate50"]); cl.alignment=aln()
        merge(ws, r,half+1,r,LAST)
        cr = ws.cell(row=r, column=half+1, value=f"  {k2}  :  {v2}" if k2 else "")
        cr.font=fnt(size=9); cr.fill=fill(C["slate50"]); cr.alignment=aln()

    # Baris 7: spasi
    ws.row_dimensions[7].height = 5

    # Baris 8-9: Header tabel
    R_H1=8; R_H2=9
    ws.row_dimensions[R_H1].height = 34
    ws.row_dimensions[R_H2].height = 28

    def hdr(r, c, val):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font=fnt(True,9,C["white"]); cell.fill=fill(C["indigo"])
        cell.alignment=aln('center', wrap=True); cell.border=thin(C["indigo2"])

    def hdr2(r, c, val):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font=fnt(True,8,C["white"]); cell.fill=fill(C["indigo2"])
        cell.alignment=aln('center', wrap=True); cell.border=thin(C["indigo3"])

    for ci, lbl in [(COL_NO,"No"), (COL_NIS,"NIS"), (COL_NAMA,"Nama Siswa")]:
        merge(ws, R_H1,ci,R_H2,ci); hdr(R_H1, ci, lbl)

    merge(ws, R_H1,COL_P1,R_H1,COL_PE)
    hdr(R_H1, COL_P1, "Pertemuan Ke-")

    for ci, lbl in [(COL_H,"Hadir\n(H)"), (COL_S,"Sakit\n(S)"), (COL_I,"Izin\n(I)"),
                    (COL_A,"Alpa\n(A)"),   (COL_TOT,"Total"),    (COL_PCT,"% Hadir")]:
        merge(ws, R_H1,ci,R_H2,ci); hdr(R_H1, ci, lbl)

    for i, pt in enumerate(pertemuan):
        hdr2(R_H2, COL_P1+i, f"P{pt['no']}\n{pt['tgl']}")

    # Data siswa
    R_DATA = R_H2 + 1
    for si, s in enumerate(siswa):
        r  = R_DATA + si
        bg = C["white"] if si % 2 == 0 else C["slate50"]
        ws.row_dimensions[r].height = 20

        sc(ws, r, COL_NO,   si+1,          h='center', bg=bg)
        sc(ws, r, COL_NIS,  str(s["nis"]), h='center', bg=bg)
        sc(ws, r, COL_NAMA, s["nama"],     bold=True,  bg=bg)

        h_=s_=i_=a_=0
        for pi, st in enumerate(s["status"]):
            ci   = COL_P1 + pi
            cell = ws.cell(row=r, column=ci, value=st)
            cell.font      = fnt(True, 9, STATUS_FG.get(st, C["gray"]))
            cell.fill      = fill(STATUS_BG.get(st, C["slate100"]))
            cell.alignment = aln('center')
            cell.border    = thin()
            if st=="H": h_+=1
            elif st=="S": s_+=1
            elif st=="I": i_+=1
            elif st=="A": a_+=1

        total = N
        pct   = h_ / total if total else 0
        p_bg  = C["emeraldL"] if pct>=.75 else (C["amberL"] if pct>=.5 else C["roseL"])
        p_fg  = C["emerald"]  if pct>=.75 else (C["amber"]  if pct>=.5 else C["rose"])

        sc(ws, r, COL_H,   h_, bold=True, fg=C["emerald"], h='center', bg=bg)
        sc(ws, r, COL_S,   s_, bold=True, fg=C["amber"],   h='center', bg=bg)
        sc(ws, r, COL_I,   i_, bold=True, fg=C["blue"],    h='center', bg=bg)
        sc(ws, r, COL_A,   a_, bold=True, fg=C["rose"],    h='center', bg=bg)
        sc(ws, r, COL_TOT, total,                           h='center', bg=bg)

        cp = ws.cell(row=r, column=COL_PCT, value=f"{round(pct*100)}%")
        cp.font=fnt(True,9,p_fg); cp.fill=fill(p_bg)
        cp.alignment=aln('center'); cp.border=thin()

    # Baris total hadir per pertemuan
    R_TOT = R_DATA + NS
    ws.row_dimensions[R_TOT].height = 22
    merge(ws, R_TOT,COL_NO,R_TOT,COL_NAMA)
    ct = ws.cell(row=R_TOT, column=COL_NO, value="TOTAL HADIR PER PERTEMUAN")
    ct.font=fnt(True,9,C["white"]); ct.fill=fill(C["indigo"])
    ct.alignment=aln('center'); ct.border=thin()

    for i in range(N):
        ci    = COL_P1 + i
        total = sum(1 for s in siswa if s["status"][i]=="H")
        c = ws.cell(row=R_TOT, column=ci, value=total)
        c.font=fnt(True,10,C["white"]); c.fill=fill(C["emerald"])
        c.alignment=aln('center'); c.border=thin()

    for ci in range(COL_H, LAST+1):
        c = ws.cell(row=R_TOT, column=ci)
        c.fill=fill(C["indigo"]); c.border=thin()

    # Freeze & print settings
    ws.freeze_panes = ws.cell(row=R_DATA, column=COL_P1)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f'1:{R_H2}'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            length    = int(self.headers.get('Content-Length', 0))
            body      = self.rfile.read(length)
            data      = json.loads(body)
            xls_bytes = build_excel(data)

            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="Rekap_Kehadiran.xlsx"')
            self.send_header('Content-Length', str(len(xls_bytes)))
            self._cors()
            self.end_headers()
            self.wfile.write(xls_bytes)
        except Exception as e:
            import traceback
            err_msg = traceback.format_exc()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self._cors()
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e), "trace": err_msg}).encode())

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin',  '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')